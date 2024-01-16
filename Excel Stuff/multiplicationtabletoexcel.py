# This program takes a number N from the command line 
# and creates an N by N multiplication table in an Excel spreadsheet.

import sys, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

# Default n is 6
if len(sys.argv) > 1:
	n = int(sys.argv[1])
else:
	n = 6

# Open a new workbook and swtich the active worksheet
wb = openpyxl.Workbook()
sheet = wb.active

# Print the headers
boldFont = Font(bold=True)
for i in range(2, n+2):
	multiplier = i - 1
	leftCell = 'A' + str(i)
	sheet[leftCell] = multiplier
	sheet[leftCell].font = boldFont
	rightCell = get_column_letter(i) + '1'
	sheet[rightCell] = multiplier
	sheet[rightCell].font = boldFont

# Print the multiplication table
for i in range(2, n+2):
	leftMultiplier = i - 1
	for j in range(2, n+2):
		rightMultiplier = j - 1
		cell = get_column_letter(j) + str(i)
		sheet[cell] = leftMultiplier*rightMultiplier

# Freeze the headers
sheet.freeze_panes = 'B2'

# Save to new Excel file in current working directory
wb.save('multiplicationTable.xlsx')